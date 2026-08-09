#!/usr/bin/env python3
"""폴더명만 보고 홍보 환자 엑셀을 일괄 백필한다 (GUI).

대상 폴더 아래 있는 세션 폴더명(``홍보여부_이름_수술명_수술날짜`` 등)을 분석해
HP/HT 환자 중 엑셀에 아직 없는 사람을 찾아 새 행으로 추가한다. 평소엔
qr_classifier.py가 세션 시작 시점에 자동으로 등록하므로, 이 도구는 과거에
누락된 폴더를 나중에 한 번에 채워 넣을 때만 쓰면 된다.
"""

from __future__ import annotations

import json
import os
import re
import sys
import threading
import tkinter as tk
from datetime import datetime
from pathlib import Path
from queue import Empty, Queue
from tkinter import filedialog, messagebox, scrolledtext, ttk

# PyInstaller onefile로 얼리면 __file__은 실행할 때마다 생기는 임시 압축 해제
# 폴더를 가리키므로, exe가 실제로 위치한 폴더를 기준으로 삼아야 설정 파일이
# 다음 실행에도 남아있는다.
if getattr(sys, "frozen", False):
    BASE_DIR = Path(sys.executable).resolve().parent
else:
    BASE_DIR = Path(__file__).resolve().parent

CONFIG_PATH = BASE_DIR / "list_updator_config.json"

DEFAULT_TARGET_DIR = Path(r"Z:\01_환자이름별사진")
DEFAULT_EXCEL_PATH = DEFAULT_TARGET_DIR / "완성형_홍보환자_리스트.xlsx"

PROMO_CHOICES = ("HP", "HT", "일반")


def load_config() -> dict:
    try:
        data = json.loads(CONFIG_PATH.read_text(encoding="utf-8"))
        return data if isinstance(data, dict) else {}
    except (FileNotFoundError, json.JSONDecodeError, OSError):
        return {}


def save_config(config: dict) -> None:
    try:
        CONFIG_PATH.write_text(json.dumps(config, ensure_ascii=False, indent=2), encoding="utf-8")
    except OSError:
        pass


def parse_folder_name(folder_name: str) -> dict | None:
    """
    폴더명을 분석하여 (구분, 이름, 수술명, 수술날짜)를 반환합니다.
    우선순위 1: 언더바(_) 구분 규칙 -> HP_홍길동_쌍꺼풀수술_260429
    우선순위 2: 띄어쓰기 혼용 규칙 -> HP 홍길동_쌍꺼풀수술_260429
    첫 조각이 HP/HT/일반 중 하나면 그 값이 구분이 되고 그 다음 조각이 이름입니다.
    첫 조각이 셋 중 무엇도 아니면 구분은 기본값 '일반'이 되고, 첫 조각이 그대로 이름이 됩니다.
    구분을 뗀 나머지가 (이름, 수술명, 날짜) 최소 3조각으로 쪼개지지 않으면 None을 반환합니다.
    """
    prefix = "일반"
    clean_name = folder_name.strip()

    for p in PROMO_CHOICES:
        if clean_name.upper().startswith(p.upper() + " ") or clean_name.upper().startswith(p.upper() + "_"):
            prefix = p
            clean_name = clean_name[len(p) + 1:].strip()
            break

    parts = [p.strip() for p in clean_name.split("_") if p.strip()]

    # 예외 처리: 만약 언더바가 아예 없다면 띄어쓰기로 분리 시도
    if len(parts) < 2:
        parts = [p.strip() for p in clean_name.split(" ") if p.strip()]

    # 이름, 수술명, 날짜 최소 3개 조각이 없으면 대상에서 제외
    if len(parts) < 3:
        return None

    date_str = None
    surgery_name = "정보 없음"
    patient_name = parts[0]

    last_part = parts[-1]
    if re.match(r"^\d{6}$", last_part):
        date_str = last_part
        if len(parts) > 2:
            surgery_name = ", ".join(parts[1:-1])
    else:
        surgery_name = ", ".join(parts[1:])

    return {"type": prefix, "name": patient_name, "surgery": surgery_name, "date_str": date_str}


def _excel_row_key(promo, name, date_value, surgery) -> tuple[str, str, str, str]:
    """A~D열(홍보여부/이름/수술날짜/수술명)로 중복 여부를 판단하기 위한 키를
    만든다. 날짜 셀은 datetime으로 저장되므로 YYMMDD 문자열로 맞춰 비교한다."""
    if isinstance(date_value, datetime):
        date_key = date_value.strftime("%y%m%d")
    else:
        date_key = str(date_value).strip() if date_value else ""
    return (
        str(promo).strip() if promo else "",
        str(name).strip() if name else "",
        date_key,
        str(surgery).strip() if surgery else "",
    )


def update_excel(target_dir: Path, excel_path: Path, log) -> int:
    """target_dir 아래 HP/HT 세션 폴더 중 excel_path에 없는 환자를 찾아 새 행으로
    추가하고, 새로 추가한 인원 수를 반환한다."""
    import openpyxl
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

    if not target_dir.is_dir():
        log(f"[오류] 대상 폴더가 존재하지 않습니다: {target_dir}")
        return 0
    if not excel_path.exists():
        log(f"[오류] 엑셀 파일이 존재하지 않습니다: {excel_path}")
        return 0

    log("폴더 스캔 중...")
    new_patients = []
    for item in os.listdir(target_dir):
        if item.startswith("_"):
            continue
        if (target_dir / item).is_dir():
            parsed = parse_folder_name(item)
            if parsed and parsed["type"] in ("HP", "HT"):
                new_patients.append(parsed)

    log(f"스캔 완료: 총 {len(new_patients)}명의 홍보 환자 후보 발견.")

    wb = openpyxl.load_workbook(excel_path)
    ws = wb.active

    # 기존에 등록된 환자 식별용 데이터 수집 (중복 등록 방지). 기준: A~D열
    # (홍보여부/이름/수술날짜/수술명)
    existing_keys = set()
    for row in range(6, ws.max_row + 1):
        name_val = ws.cell(row=row, column=2).value
        if not name_val:
            continue
        existing_keys.add(_excel_row_key(
            ws.cell(row=row, column=1).value, name_val,
            ws.cell(row=row, column=3).value, ws.cell(row=row, column=4).value,
        ))

    font_body = Font(name="Malgun Gothic", size=9)
    font_bold = Font(name="Malgun Gothic", size=9, bold=True)
    align_center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    align_left = Alignment(horizontal="left", vertical="center", wrap_text=True)
    border_side = Side(border_style="thin", color="D3D3D3")
    border_cell = Border(left=border_side, right=border_side, top=border_side, bottom=border_side)
    fill_hp = PatternFill(start_color="FFE6CC", fill_type="solid")
    fill_ht = PatternFill(start_color="E2EFDA", fill_type="solid")
    fill_zebra = PatternFill(start_color="F0F4F8", fill_type="solid")

    added_count = 0
    current_row = ws.max_row + 1

    for pat in new_patients:
        date_obj = None
        if pat["date_str"]:
            try:
                date_obj = datetime.strptime(pat["date_str"], "%y%m%d")
            except ValueError:
                pass

        pat_key = _excel_row_key(pat["type"], pat["name"], pat["date_str"], pat["surgery"])
        if pat_key in existing_keys:
            continue

        ws.row_dimensions[current_row].height = 22

        # A열: 구분 (HP / HT / 일반)
        type_cell = ws.cell(row=current_row, column=1, value=pat["type"])
        if pat["type"] == "HP":
            type_cell.fill = fill_hp
        elif pat["type"] == "HT":
            type_cell.fill = fill_ht

        # B열: 이름
        ws.cell(row=current_row, column=2, value=pat["name"]).font = font_bold

        # 차트번호/전화번호였던 C, D열은 삭제되어 수술날짜가 C열로 당겨진다.
        date_cell = ws.cell(row=current_row, column=3)
        if date_obj:
            date_cell.value = date_obj
            date_cell.number_format = "yyyy-mm-dd"
        else:
            date_cell.value = ""

        # D열: 수술명
        ws.cell(row=current_row, column=4, value=pat["surgery"])

        # E~N열: 경과일 계산 수식 및 촬영여부
        r = current_row
        ws.cell(row=r, column=5, value=f"=C{r}+7").number_format = "yyyy-mm-dd"
        ws.cell(row=r, column=6, value="")
        ws.cell(row=r, column=7, value=f"=C{r}+14").number_format = "yyyy-mm-dd"
        ws.cell(row=r, column=8, value="")
        ws.cell(row=r, column=9,
                value=f"=DATE(YEAR(C{r}), MONTH(C{r})+1, DAY(C{r}))").number_format = "yyyy-mm-dd"
        ws.cell(row=r, column=10, value="")
        ws.cell(row=r, column=11,
                value=f"=DATE(YEAR(C{r}), MONTH(C{r})+3, DAY(C{r}))").number_format = "yyyy-mm-dd"
        ws.cell(row=r, column=12, value="")
        ws.cell(row=r, column=13,
                value=f"=DATE(YEAR(C{r}), MONTH(C{r})+6, DAY(C{r}))").number_format = "yyyy-mm-dd"
        ws.cell(row=r, column=14, value="")

        # O열: 비고
        ws.cell(row=r, column=15, value="")

        is_even_row = (current_row % 2 == 0)
        for col in range(1, 16):
            cell = ws.cell(row=current_row, column=col)
            cell.border = border_cell
            if col != 2:
                cell.font = font_body
            cell.alignment = align_left if col in (4, 15) else align_center
            if col > 1 and is_even_row:
                cell.fill = fill_zebra

        current_row += 1
        added_count += 1
        existing_keys.add(pat_key)

    wb.save(excel_path)
    if added_count > 0:
        log(f"[성공] 총 {added_count}명의 새로운 홍보 환자가 엑셀에 추가되었습니다!")
    else:
        log("[안내] 새로 추가할 신규 홍보 환자 폴더가 없습니다.")
    return added_count


# ==========================================
# 화면
# ==========================================
UI_FONT = ("Malgun Gothic", 10)
UI_FONT_BOLD = ("Malgun Gothic", 10, "bold")
TITLE_FONT = ("Malgun Gothic", 15, "bold")

BG_COLOR = "#f4f6fa"
CARD_COLOR = "#ffffff"
BORDER_COLOR = "#d7dbe3"
TEXT_COLOR = "#1f2937"
MUTED_COLOR = "#6b7280"
ACCENT_COLOR = "#2f6fed"
ACCENT_HOVER = "#255ac2"
SUCCESS_COLOR = "#1a7f37"
ERROR_COLOR = "#c0392b"

ENTRY_KWARGS = dict(
    font=UI_FONT, relief="flat", highlightthickness=1,
    highlightbackground=BORDER_COLOR, highlightcolor=ACCENT_COLOR,
    bg=CARD_COLOR, fg=TEXT_COLOR, insertbackground=TEXT_COLOR,
)

INITIAL_WINDOW_SIZE = (760, 620)


class ListUpdatorApp(tk.Tk):
    def __init__(self) -> None:
        super().__init__()
        self.title("홍보 환자 엑셀 백필")
        self.resizable(True, True)

        self._config = load_config()
        self._save_after_id = None
        self.target_dir_var = tk.StringVar(value=self._config.get("target_dir") or str(DEFAULT_TARGET_DIR))
        self.excel_path_var = tk.StringVar(value=self._config.get("excel_path") or str(DEFAULT_EXCEL_PATH))
        for var in (self.target_dir_var, self.excel_path_var):
            var.trace_add("write", self._schedule_save_config)

        self._worker_thread: threading.Thread | None = None
        self._log_queue: Queue = Queue()

        self._apply_style()
        self._build_ui()
        self.minsize(*INITIAL_WINDOW_SIZE)
        self._center_window()

    def _schedule_save_config(self, *_args) -> None:
        if self._save_after_id:
            self.after_cancel(self._save_after_id)
        self._save_after_id = self.after(500, self._save_config_now)

    def _save_config_now(self) -> None:
        self._save_after_id = None
        self._config.update({
            "target_dir": self.target_dir_var.get().strip(),
            "excel_path": self.excel_path_var.get().strip(),
        })
        save_config(self._config)

    def _apply_style(self) -> None:
        self.configure(bg=BG_COLOR)
        self.option_add("*Font", UI_FONT)

        style = ttk.Style(self)
        style.theme_use("clam")
        style.configure(".", font=UI_FONT, background=BG_COLOR, foreground=TEXT_COLOR)
        style.configure("TFrame", background=BG_COLOR)
        style.configure("TLabel", background=BG_COLOR, foreground=TEXT_COLOR)
        style.configure("Header.TLabel", font=TITLE_FONT, background=BG_COLOR)
        style.configure("Muted.TLabel", background=BG_COLOR, foreground=MUTED_COLOR)
        style.configure("Success.TLabel", background=BG_COLOR, foreground=SUCCESS_COLOR)
        style.configure("Error.TLabel", background=BG_COLOR, foreground=ERROR_COLOR)

        style.configure("TLabelframe", background=BG_COLOR, bordercolor=BORDER_COLOR)
        style.configure("TLabelframe.Label", background=BG_COLOR, foreground=MUTED_COLOR,
                         font=UI_FONT_BOLD)

        style.configure("Accent.TButton", font=UI_FONT_BOLD, background=ACCENT_COLOR,
                         foreground="white", borderwidth=0, padding=(12, 10))
        style.map("Accent.TButton", background=[("active", ACCENT_HOVER), ("pressed", ACCENT_HOVER)])

        style.configure("Secondary.TButton", font=UI_FONT, background=CARD_COLOR,
                         foreground=TEXT_COLOR, bordercolor=BORDER_COLOR, padding=(10, 6))
        style.map("Secondary.TButton", background=[("active", "#eef1f6")])

    def _build_ui(self) -> None:
        self.columnconfigure(0, weight=1)
        self.rowconfigure(0, weight=1)

        outer = ttk.Frame(self, padding=24)
        outer.grid(row=0, column=0, sticky="nsew")
        outer.columnconfigure(0, weight=1)
        outer.rowconfigure(4, weight=1)

        ttk.Label(outer, text="홍보 환자 엑셀 백필", style="Header.TLabel").grid(row=0, column=0, sticky="w")
        ttk.Label(outer, text="대상 폴더의 세션 폴더명을 분석해 엑셀에 없는 HP/HT 환자를 추가합니다.",
                  style="Muted.TLabel").grid(row=1, column=0, sticky="w", pady=(2, 16))

        settings_frame = ttk.LabelFrame(outer, text="폴더 설정", padding=(16, 14))
        settings_frame.grid(row=2, column=0, sticky="ew")
        settings_frame.columnconfigure(1, weight=1)

        ttk.Label(settings_frame, text="대상 폴더").grid(row=0, column=0, sticky="e", padx=(0, 12), pady=6)
        tk.Entry(settings_frame, textvariable=self.target_dir_var, **ENTRY_KWARGS).grid(
            row=0, column=1, sticky="ew", padx=(0, 8), pady=6)
        ttk.Button(settings_frame, text="찾아보기", style="Secondary.TButton",
                   command=self._browse_target_dir).grid(row=0, column=2, pady=6)

        ttk.Label(settings_frame, text="엑셀 파일").grid(row=1, column=0, sticky="e", padx=(0, 12), pady=6)
        tk.Entry(settings_frame, textvariable=self.excel_path_var, **ENTRY_KWARGS).grid(
            row=1, column=1, sticky="ew", padx=(0, 8), pady=6)
        ttk.Button(settings_frame, text="찾아보기", style="Secondary.TButton",
                   command=self._browse_excel_file).grid(row=1, column=2, pady=6, padx=(0, 8))
        ttk.Button(settings_frame, text="열기", style="Secondary.TButton",
                   command=self._open_excel).grid(row=1, column=3, pady=6)

        control_frame = ttk.Frame(outer)
        control_frame.grid(row=3, column=0, sticky="ew", pady=(16, 8))

        self.run_btn = ttk.Button(control_frame, text="실행", style="Accent.TButton",
                                   command=self._run_update)
        self.run_btn.grid(row=0, column=0, sticky="w")

        self.status_var = tk.StringVar(value="대기 중")
        self.status_label = ttk.Label(control_frame, textvariable=self.status_var, style="Muted.TLabel")
        self.status_label.grid(row=0, column=1, sticky="w", padx=(16, 0))

        log_frame = ttk.LabelFrame(outer, text="실행 로그", padding=(10, 10))
        log_frame.grid(row=4, column=0, sticky="nsew", pady=(8, 0))
        log_frame.columnconfigure(0, weight=1)
        log_frame.rowconfigure(0, weight=1)

        self.log_text = scrolledtext.ScrolledText(
            log_frame, font=UI_FONT, wrap="word", state="disabled",
            bg=CARD_COLOR, fg=TEXT_COLOR, relief="flat", highlightthickness=1,
            highlightbackground=BORDER_COLOR,
        )
        self.log_text.grid(row=0, column=0, sticky="nsew")

    def _browse_target_dir(self) -> None:
        initial = self.target_dir_var.get().strip() or str(BASE_DIR)
        selected = filedialog.askdirectory(initialdir=initial, title="대상 폴더 선택")
        if selected:
            self.target_dir_var.set(selected)

    def _browse_excel_file(self) -> None:
        current = Path(self.excel_path_var.get().strip() or DEFAULT_EXCEL_PATH)
        selected = filedialog.askopenfilename(
            initialdir=str(current.parent) if current.parent.exists() else str(BASE_DIR),
            initialfile=current.name,
            title="엑셀 파일 선택",
            filetypes=[("Excel 파일", "*.xlsx"), ("모든 파일", "*.*")],
        )
        if selected:
            self.excel_path_var.set(selected)

    def _open_excel(self) -> None:
        excel_path = Path(self.excel_path_var.get().strip() or DEFAULT_EXCEL_PATH)
        if not excel_path.exists():
            messagebox.showerror("파일 없음", f"엑셀 파일을 찾을 수 없습니다:\n{excel_path}")
            return
        try:
            os.startfile(str(excel_path))
        except OSError as exc:
            messagebox.showerror("파일 열기 실패", str(exc))

    def _center_window(self) -> None:
        self.update_idletasks()
        width, height = INITIAL_WINDOW_SIZE
        x = (self.winfo_screenwidth() - width) // 2
        y = (self.winfo_screenheight() - height) // 3
        self.geometry(f"{width}x{height}+{x}+{y}")

    def _set_status(self, text: str, *, tone: str = "muted") -> None:
        self.status_var.set(text)
        style_by_tone = {"muted": "Muted.TLabel", "success": "Success.TLabel", "error": "Error.TLabel"}
        self.status_label.configure(style=style_by_tone[tone])

    def _run_update(self) -> None:
        if self._worker_thread and self._worker_thread.is_alive():
            return

        target_dir = self.target_dir_var.get().strip()
        excel_path = self.excel_path_var.get().strip()
        if not (target_dir and excel_path):
            messagebox.showerror("입력 오류", "대상 폴더와 엑셀 파일을 모두 지정하세요.")
            return

        self.log_text.configure(state="normal")
        self.log_text.delete("1.0", "end")
        self.log_text.configure(state="disabled")

        self._log_queue = Queue()
        self.run_btn.configure(state="disabled")
        self._set_status("실행 중", tone="success")

        self._worker_thread = threading.Thread(
            target=self._worker,
            args=(Path(target_dir), Path(excel_path), self._log_queue.put),
            daemon=True,
        )
        self._worker_thread.start()
        self._poll_log_queue()

    def _worker(self, target_dir: Path, excel_path: Path, log_put) -> None:
        try:
            update_excel(target_dir, excel_path, log_put)
        except Exception as error:
            log_put(f"[오류] 실행 중 문제가 발생했습니다: {error}")

    def _poll_log_queue(self) -> None:
        while True:
            try:
                message = self._log_queue.get_nowait()
            except Empty:
                break
            self._append_log(message)

        if self._worker_thread and self._worker_thread.is_alive():
            self.after(200, self._poll_log_queue)
        else:
            self.run_btn.configure(state="normal")
            self._set_status("완료", tone="muted")

    def _append_log(self, message: str) -> None:
        self.log_text.configure(state="normal")
        self.log_text.insert("end", message + "\n")
        self.log_text.see("end")
        self.log_text.configure(state="disabled")


def main() -> None:
    app = ListUpdatorApp()
    app.mainloop()


if __name__ == "__main__":
    main()
