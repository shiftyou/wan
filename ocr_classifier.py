#!/usr/bin/env python3
"""QR 세션으로 촬영 사진을 자동 분류한다.

99_사진작업폴더\\photo(WATCH_DIR, 기본값: 이 스크립트가 있는 폴더의 photo 하위폴더)에
EOS Utility가 새 사진을 저장하면 다음 규칙으로 01_환자이름별사진(PHOTO_DIR) 아래로 옮긴다.
* ``홍보여부_이름_수술명_수술날짜`` QR 사진: 해당 촬영 세션을 시작하고 이전 세션을 끝낸다.
* QR이 없는 사진: 현재 세션의 폴더로 이동한다.
* 세션이 없을 때 들어온 사진 또는 알 수 없는 QR: ``_미분류``으로 이동한다.

QR에는 환자 정보가 들어가며, 사진은 PHOTO_DIR의 ``홍보여부_이름_수술명_수술날짜``
세션 폴더 아래 ``홍보여부_이름_경과일`` 서브폴더(수술날짜 기준 촬영일 경과일수)에
저장된다.

분류/이동 전에 원본 사진은 항상 02_날짜별사진(PHOTO_BACKUP_DIR)의
``년/년월/년월일/`` 아래에 원본 파일명 그대로 복사되어 보관된다.
"""

from __future__ import annotations

import json
import os
import re
import shutil
import sys
import time
from datetime import datetime
from pathlib import Path


BASE_DIR = Path(__file__).resolve().parent
WATCH_DIR = Path(os.environ.get("WATCH_DIR", BASE_DIR / "photo")).expanduser()
PHOTO_DIR = Path(os.environ.get("PHOTO_DIR", r"Z:\01_환자이름별사진")).expanduser()
PHOTO_BACKUP_DIR = Path(os.environ.get("PHOTO_BACKUP_DIR", r"Z:\02_날짜별사진")).expanduser()
STATE_PATH = BASE_DIR / "current_session.json"
EXCEL_PATH = PHOTO_DIR / "완성형_홍보환자_리스트.xlsx"
CHECK_INTERVAL = float(os.environ.get("CHECK_INTERVAL", "1"))
STABLE_SECONDS = float(os.environ.get("STABLE_SECONDS", "3"))
IMAGE_EXTENSIONS = {".jpg", ".jpeg", ".png", ".bmp", ".tiff", ".tif", ".webp"}

PROMO_CHOICES = ("HP", "HT", "일반")
SESSION_DATE_RE = re.compile(r"^\d{6}$")


def parse_session_payload(payload: str) -> dict[str, str] | None:
    """``[홍보여부_]이름_수술명_수술날짜`` 형식의 QR 내용을 해석한다.
    홍보여부가 생략된 3조각 payload는 '일반'으로 간주한다."""
    parts = payload.split("_")
    if len(parts) == 4:
        promo, name, surgery, surgery_date = (part.strip() for part in parts)
        if promo not in PROMO_CHOICES:
            return None
    elif len(parts) == 3:
        promo = "일반"
        name, surgery, surgery_date = (part.strip() for part in parts)
    else:
        return None
    if not name or not surgery:
        return None
    if not SESSION_DATE_RE.fullmatch(surgery_date):
        return None
    return {"홍보여부": promo, "이름": name, "수술명": surgery, "수술날짜": surgery_date}


def load_current_session() -> dict | None:
    try:
        data = json.loads(STATE_PATH.read_text(encoding="utf-8"))
    except (FileNotFoundError, json.JSONDecodeError):
        return None
    folder = data.get("folder")
    info = data.get("info")
    if isinstance(folder, str) and isinstance(info, dict):
        return {"folder": folder, "info": info}
    return None


def save_current_session(session: dict | None) -> None:
    temporary = STATE_PATH.with_suffix(".tmp")
    temporary.write_text(
        json.dumps(session or {}, ensure_ascii=False), encoding="utf-8"
    )
    os.replace(temporary, STATE_PATH)


def folder_name(value: str) -> str:
    """문자열을 사람이 읽을 수 있으면서 파일 경로로 안전한 폴더명으로 바꾼다."""
    clean_value = re.sub(r'[<>:"/\\|?*\n\r\t]', " ", value).strip()
    clean_value = re.sub(r"\s+", " ", clean_value)
    if not clean_value:
        raise ValueError("폴더명이 비어 있습니다")
    return clean_value[:80].rstrip()


def build_session_folder(info: dict[str, str]) -> str:
    """QR 페이로드 형식([홍보여부_]이름_수술명_수술날짜)을 그대로 세션 폴더명으로 쓴다.
    QR과 동일하게 홍보여부가 '일반'이면 표기하지 않는다."""
    if info["홍보여부"] == "일반":
        raw = f"{info['이름']}_{info['수술명']}_{info['수술날짜']}"
    else:
        raw = f"{info['홍보여부']}_{info['이름']}_{info['수술명']}_{info['수술날짜']}"
    return folder_name(raw)


def build_elapsed_subfolder(info: dict[str, str], image_path: Path) -> str:
    """수술날짜 기준으로 사진이 찍힌 날짜까지의 경과일 서브폴더명을 만든다.
    QR과 동일하게 홍보여부가 '일반'이면 표기하지 않는다."""
    photo_date = datetime.fromtimestamp(image_path.stat().st_mtime).date()
    surgery_date = datetime.strptime(info["수술날짜"], "%y%m%d").date()
    elapsed_days = (photo_date - surgery_date).days
    if info["홍보여부"] == "일반":
        raw = f"{info['이름']}_#{elapsed_days}"
    else:
        raw = f"{info['홍보여부']}_{info['이름']}_#{elapsed_days}"
    return folder_name(raw)


def safe_move(source: Path, destination_dir: Path) -> Path:
    destination_dir.mkdir(parents=True, exist_ok=True)
    destination = destination_dir / source.name
    if destination.exists():
        stem, suffix = source.stem, source.suffix
        index = 1
        while destination.exists():
            destination = destination_dir / f"{stem}_{index}{suffix}"
            index += 1
    shutil.move(str(source), str(destination))
    return destination


def backup_original(image_path: Path) -> None:
    """분류/이동 전에 원본 사진을 02_날짜별사진/YYYY/YYYY-MM/YYYY-MM-DD/ 아래에 복사해 둔다."""
    photo_date = datetime.fromtimestamp(image_path.stat().st_mtime)
    backup_dir = (
        PHOTO_BACKUP_DIR
        / photo_date.strftime("%Y")
        / photo_date.strftime("%Y-%m")
        / photo_date.strftime("%Y-%m-%d")
    )
    backup_dir.mkdir(parents=True, exist_ok=True)
    destination = backup_dir / image_path.name
    if destination.exists():
        stem, suffix = image_path.stem, image_path.suffix
        index = 1
        while destination.exists():
            destination = backup_dir / f"{stem}_{index}{suffix}"
            index += 1
    shutil.copy2(image_path, destination)


def move_with_sequence(source: Path, destination_dir: Path, base_name: str) -> Path:
    """대상 폴더명 + 순번으로 파일명을 바꿔서 옮긴다 (예: 일반_이름_#7_1.jpg)."""
    destination_dir.mkdir(parents=True, exist_ok=True)
    index = 1
    if destination_dir.exists():
        for item in destination_dir.iterdir():
            match = re.fullmatch(rf"{re.escape(base_name)}_(\d+)", item.stem)
            if match:
                index = max(index, int(match.group(1)) + 1)
    destination = destination_dir / f"{base_name}_{index}{source.suffix}"
    while destination.exists():
        index += 1
        destination = destination_dir / f"{base_name}_{index}{source.suffix}"
    shutil.move(str(source), str(destination))
    return destination


def _detect_qr(detector, image) -> str | None:
    data, _points, _straight = detector.detectAndDecode(image)
    return data.strip() if data else None


def _detect_qr_in_tiles(detector, image, grid: int = 2, overlap: float = 0.15) -> str | None:
    """QR이 사진 전체에 비해 작게 찍혔을 때를 대비해, 원본 해상도 그대로
    구역을 나눠 다시 검사한다. detectAndDecode는 큰 이미지를 내부적으로
    축소해서 처리하기 때문에, 전체 사진 기준으로 QR이 작으면 인식에
    실패하기 쉽다 - 구역별로 잘라서 보면 QR이 상대적으로 커져 인식이
    잘 된다."""
    height, width = image.shape[:2]
    tile_h, tile_w = height // grid, width // grid
    pad_h, pad_w = int(tile_h * overlap), int(tile_w * overlap)
    for row in range(grid):
        for col in range(grid):
            y0 = max(0, row * tile_h - pad_h)
            y1 = min(height, (row + 1) * tile_h + pad_h)
            x0 = max(0, col * tile_w - pad_w)
            x1 = min(width, (col + 1) * tile_w + pad_w)
            payload = _detect_qr(detector, image[y0:y1, x0:x1])
            if payload:
                return payload
    return None


def decode_qr(detector, image_path: Path) -> str | None:
    """사진 한 장에서 하나의 QR payload를 읽는다. 인식은 모두 로컬에서 수행된다."""
    import cv2
    import numpy as np

    # cv2.imread는 Windows에서 한글(비 ASCII)이 섞인 경로, 특히 UNC 네트워크
    # 경로를 제대로 열지 못한다 (내부적으로 ANSI 코드페이지 API를 사용).
    # 파일을 직접 바이트로 읽어 imdecode에 넘기면 경로 문제를 피할 수 있다.
    try:
        file_bytes = np.fromfile(str(image_path), dtype=np.uint8)
    except OSError:
        return None
    if file_bytes.size == 0:
        return None
    image = cv2.imdecode(file_bytes, cv2.IMREAD_COLOR)
    if image is None:
        return None

    payload = _detect_qr(detector, image)
    if payload:
        return payload
    return _detect_qr_in_tiles(detector, image)


def incoming_images() -> list[Path]:
    return sorted(
        (
            item
            for item in WATCH_DIR.iterdir()
            if item.is_file() and item.suffix.lower() in IMAGE_EXTENSIONS
        ),
        key=lambda item: item.stat().st_mtime,
    )


def is_stable(image_path: Path, observations: dict[Path, tuple[int, int, float]]) -> bool:
    """EOS Utility가 쓰는 중인 파일을 처리하지 않는다."""
    try:
        stat = image_path.stat()
    except FileNotFoundError:
        observations.pop(image_path, None)
        return False

    signature = (stat.st_size, stat.st_mtime_ns)
    previous = observations.get(image_path)
    now = time.monotonic()
    if previous is None or previous[:2] != signature:
        observations[image_path] = (*signature, now)
        return False
    return now - previous[2] >= STABLE_SECONDS


def write_info(folder: Path, info: dict[str, str]) -> None:
    folder.mkdir(parents=True, exist_ok=True)
    (folder / "info.json").write_text(
        json.dumps(info, ensure_ascii=False, indent=2), encoding="utf-8"
    )


def append_excel_row(info: dict[str, str]) -> None:
    """새 세션이 시작될 때 홍보 환자 엑셀에 한 행을 바로 추가한다."""
    try:
        import openpyxl
        from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    except ImportError as error:
        print(f"  엑셀 업데이트 건너뜀 (openpyxl 필요): {error}")
        return

    if not EXCEL_PATH.exists():
        print(f"  엑셀 업데이트 건너뜀 (파일 없음): {EXCEL_PATH}")
        return

    try:
        wb = openpyxl.load_workbook(EXCEL_PATH)
        ws = wb.active

        existing_names = set()
        for row in range(6, ws.max_row + 1):
            name_val = ws.cell(row=row, column=2).value
            if name_val:
                existing_names.add(str(name_val).strip())

        if info["이름"] in existing_names:
            print(f"  엑셀 업데이트 건너뜀 (이미 등록됨): {info['이름']}")
            return

        try:
            surgery_date = datetime.strptime(info["수술날짜"], "%y%m%d")
        except ValueError:
            surgery_date = None

        font_body = Font(name="Malgun Gothic", size=9)
        font_bold = Font(name="Malgun Gothic", size=9, bold=True)
        align_center = Alignment(horizontal="center", vertical="center", wrap_text=True)
        align_left = Alignment(horizontal="left", vertical="center", wrap_text=True)
        border_side = Side(border_style="thin", color="D3D3D3")
        border_cell = Border(left=border_side, right=border_side, top=border_side, bottom=border_side)
        fill_hp = PatternFill(start_color="FFE6CC", fill_type="solid")
        fill_ht = PatternFill(start_color="E2EFDA", fill_type="solid")
        fill_zebra = PatternFill(start_color="F0F4F8", fill_type="solid")

        row = ws.max_row + 1
        ws.row_dimensions[row].height = 22

        type_cell = ws.cell(row=row, column=1, value=info["홍보여부"])
        if info["홍보여부"] == "HP":
            type_cell.fill = fill_hp
        elif info["홍보여부"] == "HT":
            type_cell.fill = fill_ht

        ws.cell(row=row, column=2, value=info["이름"]).font = font_bold
        ws.cell(row=row, column=3, value="")
        ws.cell(row=row, column=4, value="")

        date_cell = ws.cell(row=row, column=5)
        if surgery_date:
            date_cell.value = surgery_date
            date_cell.number_format = "yyyy-mm-dd"
        else:
            date_cell.value = ""

        ws.cell(row=row, column=6, value=info["수술명"])

        ws.cell(row=row, column=7, value=f"=E{row}+7").number_format = "yyyy-mm-dd"
        ws.cell(row=row, column=8, value="")
        ws.cell(row=row, column=9, value=f"=E{row}+14").number_format = "yyyy-mm-dd"
        ws.cell(row=row, column=10, value="")
        ws.cell(row=row, column=11,
                value=f"=DATE(YEAR(E{row}), MONTH(E{row})+1, DAY(E{row}))").number_format = "yyyy-mm-dd"
        ws.cell(row=row, column=12, value="")
        ws.cell(row=row, column=13,
                value=f"=DATE(YEAR(E{row}), MONTH(E{row})+3, DAY(E{row}))").number_format = "yyyy-mm-dd"
        ws.cell(row=row, column=14, value="")
        ws.cell(row=row, column=15,
                value=f"=DATE(YEAR(E{row}), MONTH(E{row})+6, DAY(E{row}))").number_format = "yyyy-mm-dd"
        ws.cell(row=row, column=16, value="")
        ws.cell(row=row, column=17, value="")

        is_even_row = (row % 2 == 0)
        for col in range(1, 18):
            cell = ws.cell(row=row, column=col)
            cell.border = border_cell
            if col != 2:
                cell.font = font_body
            cell.alignment = align_left if col in (6, 17) else align_center
            if col > 1 and is_even_row:
                cell.fill = fill_zebra

        wb.save(EXCEL_PATH)
        print(f"  엑셀 업데이트: {info['이름']} 추가")
    except Exception as error:
        print(f"  엑셀 업데이트 실패: {error}")


def process_image(detector, image_path: Path, current_session: dict | None) -> dict | None:
    backup_original(image_path)
    payload = decode_qr(detector, image_path)

    if payload:
        info = parse_session_payload(payload)
        if info:
            session_folder = build_session_folder(info)
            marker_dir = PHOTO_DIR / "_세션마커" / session_folder
            safe_move(image_path, marker_dir)
            write_info(marker_dir, info)
            write_info(PHOTO_DIR / session_folder, info)
            if current_session and current_session["folder"] != session_folder:
                print(f"  이전 세션 종료: {current_session['folder']}")
            print(f"  세션 시작: {session_folder}")
            append_excel_row(info)
            return {"folder": session_folder, "info": info}

        safe_move(image_path, PHOTO_DIR / "_미분류" / "알수없는_QR")
        print("  형식이 올바르지 않은 QR -> _미분류")
        return current_session

    if current_session:
        subfolder = build_elapsed_subfolder(current_session["info"], image_path)
        destination = move_with_sequence(
            image_path, PHOTO_DIR / current_session["folder"] / subfolder, subfolder
        )
        print(f"  사진 저장: {current_session['folder']}/{subfolder}/{destination.name}")
        return current_session

    safe_move(image_path, PHOTO_DIR / "_미분류")
    print("  시작 QR 전 사진 -> _미분류/")
    return None


def ensure_packages() -> None:
    try:
        import cv2  # noqa: F401
    except ImportError:
        print("OpenCV가 필요합니다. 아래 명령을 한 번 실행하세요:\n"
              "  .venv/bin/python -m pip install opencv-python")
        sys.exit(1)


def wait_for_directory(path: Path, retry_seconds: float = 5.0) -> None:
    """네트워크 드라이브(Z: 등)가 로그인 직후 아직 마운트되지 않았을 수 있어 재시도한다."""
    while True:
        try:
            path.mkdir(parents=True, exist_ok=True)
            return
        except OSError as error:
            print(f"  {path} 접근 대기 중... ({error})")
            time.sleep(retry_seconds)


def main() -> None:
    ensure_packages()
    import cv2

    wait_for_directory(WATCH_DIR)
    wait_for_directory(PHOTO_DIR)
    current_session = load_current_session()

    print("=" * 58)
    print("  QR 촬영 세션 자동 분류기")
    print(f"  감시 폴더: {WATCH_DIR}")
    print(f"  분류 대상 폴더: {PHOTO_DIR}")
    print(f"  파일 안정화: {STABLE_SECONDS:g}초")
    print(f"  현재 세션: {current_session['folder'] if current_session else '없음'}")
    print("  종료: Ctrl+C")
    print("=" * 58)

    detector = cv2.QRCodeDetector()
    observations: dict[Path, tuple[int, int, float]] = {}

    while True:
        for image_path in incoming_images():
            if not is_stable(image_path, observations):
                continue
            print(f"\n[처리] {image_path.name}")
            try:
                current_session = process_image(detector, image_path, current_session)
                save_current_session(current_session)
                observations.pop(image_path, None)
            except Exception as error:
                # 파일을 남겨 다음 주기에 다시 시도한다. 실패 파일을 처리 완료로 표시하지 않는다.
                print(f"  처리 보류 (다시 시도): {error}")
        time.sleep(CHECK_INTERVAL)


if __name__ == "__main__":
    try:
        main()
    except KeyboardInterrupt:
        print("\n분류기를 종료합니다.")
