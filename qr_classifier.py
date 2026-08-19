#!/usr/bin/env python3
"""QR 세션으로 촬영 사진을 자동 분류한다 (GUI + 백그라운드 감시).

감시 폴더에 EOS Utility가 새 사진을 저장하면 다음 규칙으로 분류 대상 폴더
아래로 옮긴다.
* ``홍보여부_이름_수술명_수술날짜`` QR 사진: 해당 촬영 세션을 시작하고 이전 세션을 끝낸다.
* QR이 없는 사진: 현재 세션의 폴더로 이동한다.
* 세션이 없을 때 들어온 사진 또는 알 수 없는 QR: ``_미분류``으로 이동한다.

QR에는 환자 정보가 들어가며, 사진은 분류 대상 폴더의 ``홍보여부_이름`` (일반은 ``이름``)
상위 폴더 아래 ``홍보여부_이름_수술명_수술날짜`` 세션 폴더, 그 아래 다시
``홍보여부_이름_수술명_경과일`` 서브폴더(수술날짜 기준 촬영일 경과일수)에
``홍보여부_이름_수술명_경과일_순번`` 파일명으로 저장된다.

분류/이동 전에 원본 사진은 항상 원본 백업 폴더의 ``년/년월/년월일/`` 아래에도
그대로 보관된다. 이 날짜 폴더 아래에는 분류 대상 폴더와 동일하게
``홍보구분_이름/세션 폴더/경과일 서브폴더`` 구조 및 파일명으로도 다시 정리되어
저장된다(세션을 시작하는 QR 사진은 원본 파일명 그대로 세션 폴더까지만).

감시 폴더/분류 대상 폴더/원본 백업 폴더는 화면에서 바로 바꿀 수 있고, 마지막으로
쓴 값이 qr_classifier_config.json에 저장되어 다음 실행 때도 불러온다.
"""

from __future__ import annotations

import json
import os
import re
import shutil
import sys
import threading
import time
import tkinter as tk
from datetime import datetime
from pathlib import Path
from queue import Empty, Queue
from tkinter import filedialog, messagebox, scrolledtext, ttk


# PyInstaller onefile로 얼리면 __file__은 실행할 때마다 생기는 임시 압축 해제
# 폴더를 가리키므로, exe가 실제로 위치한 폴더를 기준으로 삼아야 설정/상태 파일이
# 다음 실행에도 남아있는다.
if getattr(sys, "frozen", False):
    BASE_DIR = Path(sys.executable).resolve().parent
else:
    BASE_DIR = Path(__file__).resolve().parent

CONFIG_PATH = BASE_DIR / "qr_classifier_config.json"
STATE_PATH = BASE_DIR / "current_session.json"

DEFAULT_WATCH_DIR = BASE_DIR / "photo"
DEFAULT_PHOTO_DIR = Path(r"Z:\01_환자이름별사진")
DEFAULT_BACKUP_DIR = Path(r"Z:\02_날짜별사진")
DEFAULT_EXCEL_PATH = DEFAULT_PHOTO_DIR / "완성형_홍보환자_리스트.xlsx"

CHECK_INTERVAL = float(os.environ.get("CHECK_INTERVAL", "1"))
STABLE_SECONDS = float(os.environ.get("STABLE_SECONDS", "3"))
IMAGE_EXTENSIONS = {".jpg", ".jpeg", ".png", ".bmp", ".tiff", ".tif", ".webp"}

# QR 검출 전 이미지를 이 크기(긴 변 기준, px) 이하로 축소한다. QR 인식에는 원본
# 해상도(수천만 화소)가 전혀 필요 없는데, cv2/pyzbar 검출기는 픽셀 수에 비례해
# 느려진다 - 특히 QR이 없는 일반 사진마다 매번 여러 번(전체+타일) 검사를 반복하므로
# 이 축소가 없으면 사진 한 장 처리 시간이 크게 늘어난다.
DECODE_MAX_DIMENSION = 1600

PROMO_CHOICES = ("HP", "HT", "일반")
SESSION_DATE_RE = re.compile(r"^\d{6}$")


def load_config() -> dict:
    try:
        data = json.loads(CONFIG_PATH.read_text(encoding="utf-8"))
        return data if isinstance(data, dict) else {}
    except (FileNotFoundError, json.JSONDecodeError, OSError):
        return {}


def save_config(config: dict) -> None:
    try:
        CONFIG_PATH.write_text(json.dumps(config, ensure_ascii=False, indent=2), encoding="utf-8")
    except OSError:
        pass


# ==========================================
# 분류 로직 (경로는 인자로 받는다 - 화면에서 바꿀 수 있으므로 전역 상수로 고정하지 않는다)
# ==========================================
def parse_session_payload(payload: str) -> dict[str, str] | None:
    """``[홍보여부_]이름_수술명_수술날짜`` 형식의 QR 내용을 해석한다.
    홍보여부가 생략된 3조각 payload는 '일반'으로 간주한다."""
    parts = payload.split("_")
    if len(parts) == 4:
        promo, name, surgery, surgery_date = (part.strip() for part in parts)
        if promo not in PROMO_CHOICES:
            return None
    elif len(parts) == 3:
        promo = "일반"
        name, surgery, surgery_date = (part.strip() for part in parts)
    else:
        return None
    if not name or not surgery:
        return None
    if not SESSION_DATE_RE.fullmatch(surgery_date):
        return None
    return {"홍보여부": promo, "이름": name, "수술명": surgery, "수술날짜": surgery_date}


def load_current_session() -> dict | None:
    try:
        data = json.loads(STATE_PATH.read_text(encoding="utf-8"))
    except (FileNotFoundError, json.JSONDecodeError):
        return None
    folder = data.get("folder")
    info = data.get("info")
    if isinstance(folder, str) and isinstance(info, dict):
        return {"folder": folder, "info": info}
    return None


def save_current_session(session: dict | None) -> None:
    temporary = STATE_PATH.with_suffix(".tmp")
    temporary.write_text(
        json.dumps(session or {}, ensure_ascii=False), encoding="utf-8"
    )
    os.replace(temporary, STATE_PATH)


def folder_name(value: str) -> str:
    """문자열을 사람이 읽을 수 있으면서 파일 경로로 안전한 폴더명으로 바꾼다."""
    clean_value = re.sub(r'[<>:"/\\|?*\n\r\t]', " ", value).strip()
    clean_value = re.sub(r"\s+", " ", clean_value)
    if not clean_value:
        raise ValueError("폴더명이 비어 있습니다")
    return clean_value[:80].rstrip()


def build_session_folder(info: dict[str, str]) -> str:
    """QR 페이로드 형식([홍보여부_]이름_수술명_수술날짜)을 그대로 세션 폴더명으로 쓴다.
    QR과 동일하게 홍보여부가 '일반'이면 표기하지 않는다."""
    if info["홍보여부"] == "일반":
        raw = f"{info['이름']}_{info['수술명']}_{info['수술날짜']}"
    else:
        raw = f"{info['홍보여부']}_{info['이름']}_{info['수술명']}_{info['수술날짜']}"
    return folder_name(raw)


def build_session_parent_folder(info: dict[str, str]) -> str:
    """세션 폴더를 담는 상위 폴더명(이름, 또는 홍보구분_이름)을 만든다.
    같은 환자의 여러 세션 폴더를 한 상위 폴더 아래 모아 두기 위함이다."""
    if info["홍보여부"] == "일반":
        raw = info["이름"]
    else:
        raw = f"{info['홍보여부']}_{info['이름']}"
    return folder_name(raw)


def build_elapsed_subfolder(info: dict[str, str], image_path: Path) -> str:
    """수술날짜 기준으로 사진이 찍힌 날짜까지의 경과일 서브폴더명을 만든다.
    이 이름은 그 안에 저장되는 사진 파일명의 접두어(+ 순번)로도 그대로 쓰인다.
    상위 폴더명에 이미 수술명이 들어가 있으므로 여기서는 뺀다.
    QR과 동일하게 홍보여부가 '일반'이면 표기하지 않는다."""
    photo_date = datetime.fromtimestamp(image_path.stat().st_mtime).date()
    surgery_date = datetime.strptime(info["수술날짜"], "%y%m%d").date()
    elapsed_days = (photo_date - surgery_date).days
    if info["홍보여부"] == "일반":
        raw = f"{info['이름']}_#{elapsed_days}"
    else:
        raw = f"{info['홍보여부']}_{info['이름']}_#{elapsed_days}"
    return folder_name(raw)


def safe_move(source: Path, destination_dir: Path) -> Path:
    destination_dir.mkdir(parents=True, exist_ok=True)
    destination = destination_dir / source.name
    if destination.exists():
        stem, suffix = source.stem, source.suffix
        index = 1
        while destination.exists():
            destination = destination_dir / f"{stem}_{index}{suffix}"
            index += 1
    shutil.move(str(source), str(destination))
    return destination


def backup_original(image_path: Path, backup_dir: Path, relative_subdir: Path | None = None,
                     filename: str | None = None) -> None:
    """분류/이동 전에 원본 사진을 backup_dir/YYYY/YYYY-MM/YYYY-MM-DD/[relative_subdir/] 아래에 복사해 둔다.
    relative_subdir을 주면 분류 대상 폴더(photo_dir)에 실제로 저장되는 홍보구분_이름/세션폴더/경과일폴더
    구조와 동일하게 날짜 폴더 아래에도 그대로 맞춘다. filename을 주면 원본 파일명 대신 그 이름을
    쓴다 (분류 대상 폴더에 실제로 저장되는 순번 파일명과 맞추기 위함)."""
    photo_date = datetime.fromtimestamp(image_path.stat().st_mtime)
    dated_dir = (
        backup_dir
        / photo_date.strftime("%Y")
        / photo_date.strftime("%Y-%m")
        / photo_date.strftime("%Y-%m-%d")
    )
    if relative_subdir:
        dated_dir = dated_dir / relative_subdir
    dated_dir.mkdir(parents=True, exist_ok=True)
    name = filename or image_path.name
    destination = dated_dir / name
    if destination.exists():
        stem, suffix = Path(name).stem, Path(name).suffix
        index = 1
        while destination.exists():
            destination = dated_dir / f"{stem}_{index}{suffix}"
            index += 1
    shutil.copy2(image_path, destination)


def next_sequence_filename(destination_dir: Path, base_name: str, suffix: str) -> str:
    """destination_dir 안에 이미 있는 ``base_name_숫자`` 꼴 파일들을 보고 다음 순번
    파일명을 만든다 (예: 일반_이름_수술명_#7_1.jpg)."""
    index = 1
    if destination_dir.exists():
        for item in destination_dir.iterdir():
            match = re.fullmatch(rf"{re.escape(base_name)}_(\d+)", item.stem)
            if match:
                index = max(index, int(match.group(1)) + 1)
    filename = f"{base_name}_{index}{suffix}"
    while (destination_dir / filename).exists():
        index += 1
        filename = f"{base_name}_{index}{suffix}"
    return filename


def _detect_qr(detector, image) -> str | None:
    data, _points, _straight = detector.detectAndDecode(image)
    return data.strip() if data else None


def _detect_qr_pyzbar(image) -> str | None:
    """cv2.QRCodeDetector는 고전적인(비 딥러닝) 알고리즘이라 블러/기울기/저대비
    상황에서 인식에 실패하는 경우가 있다. ZBar 기반 pyzbar로 한 번 더 시도해
    인식률을 보완한다. pyzbar가 설치되어 있지 않거나(ImportError) 네이티브 zbar
    DLL을 못 찾으면(OSError - 특히 exe로 얼렸을 때) 조용히 건너뛴다. 이 함수가
    예외를 흘리면 호출부에서 사진 처리 자체가 계속 실패해 버리므로 반드시
    잡아야 한다."""
    try:
        from pyzbar.pyzbar import ZBarSymbol
        from pyzbar.pyzbar import decode as zbar_decode
    except (ImportError, OSError):
        return None

    try:
        import cv2

        gray = cv2.cvtColor(image, cv2.COLOR_BGR2GRAY)
        results = zbar_decode(gray, symbols=[ZBarSymbol.QRCODE])
    except OSError:
        return None
    if not results:
        return None
    return results[0].data.decode("utf-8", errors="replace").strip() or None


def _scan_tiles(image, detect_fn, grid: int = 2, overlap: float = 0.15) -> str | None:
    """QR이 사진 전체에 비해 작게 찍혔을 때를 대비해, 원본 해상도 그대로
    구역을 나눠 detect_fn으로 다시 검사한다. 전체 이미지 기준 검출기는 내부적으로
    이미지를 축소해서 처리하는 경우가 많아, 전체 사진 기준으로 QR이 작으면 인식에
    실패하기 쉽다 - 구역별로 잘라서 보면 QR이 상대적으로 커져 인식이 잘 된다."""
    height, width = image.shape[:2]
    tile_h, tile_w = height // grid, width // grid
    pad_h, pad_w = int(tile_h * overlap), int(tile_w * overlap)
    for row in range(grid):
        for col in range(grid):
            y0 = max(0, row * tile_h - pad_h)
            y1 = min(height, (row + 1) * tile_h + pad_h)
            x0 = max(0, col * tile_w - pad_w)
            x1 = min(width, (col + 1) * tile_w + pad_w)
            payload = detect_fn(image[y0:y1, x0:x1])
            if payload:
                return payload
    return None


def decode_qr(detector, image_path: Path, use_pyzbar: bool = False) -> str | None:
    """사진 한 장에서 하나의 QR payload를 읽는다. 인식은 모두 로컬에서 수행된다.
    cv2(전체 -> 구역별)로 먼저 시도하고, use_pyzbar가 True면 실패했을 때 pyzbar
    (전체 -> 구역별)로 한 번 더 시도한다. QR 인식용으로만 쓰이는 복사본이라
    DECODE_MAX_DIMENSION으로 미리 축소해서 검사한다 - 실제로 이동/백업되는
    원본 파일에는 영향이 없다."""
    import cv2
    import numpy as np

    # cv2.imread는 Windows에서 한글(비 ASCII)이 섞인 경로, 특히 UNC 네트워크
    # 경로를 제대로 열지 못한다 (내부적으로 ANSI 코드페이지 API를 사용).
    # 파일을 직접 바이트로 읽어 imdecode에 넘기면 경로 문제를 피할 수 있다.
    try:
        file_bytes = np.fromfile(str(image_path), dtype=np.uint8)
    except OSError:
        return None
    if file_bytes.size == 0:
        return None
    image = cv2.imdecode(file_bytes, cv2.IMREAD_COLOR)
    if image is None:
        return None

    height, width = image.shape[:2]
    longest_side = max(height, width)
    if longest_side > DECODE_MAX_DIMENSION:
        scale = DECODE_MAX_DIMENSION / longest_side
        image = cv2.resize(image, (int(width * scale), int(height * scale)),
                            interpolation=cv2.INTER_AREA)

    detect_fns = [lambda img: _detect_qr(detector, img)]
    if use_pyzbar:
        detect_fns.append(_detect_qr_pyzbar)

    for detect_fn in detect_fns:
        payload = detect_fn(image)
        if payload:
            return payload
        payload = _scan_tiles(image, detect_fn)
        if payload:
            return payload
    return None


def incoming_images(watch_dir: Path) -> list[Path]:
    return sorted(
        (
            item
            for item in watch_dir.iterdir()
            if item.is_file() and item.suffix.lower() in IMAGE_EXTENSIONS
        ),
        key=lambda item: item.stat().st_mtime,
    )


def is_stable(image_path: Path, observations: dict[Path, tuple[int, int, float]]) -> bool:
    """EOS Utility가 쓰는 중인 파일을 처리하지 않는다."""
    try:
        stat = image_path.stat()
    except FileNotFoundError:
        observations.pop(image_path, None)
        return False

    signature = (stat.st_size, stat.st_mtime_ns)
    previous = observations.get(image_path)
    now = time.monotonic()
    if previous is None or previous[:2] != signature:
        observations[image_path] = (*signature, now)
        return False
    return now - previous[2] >= STABLE_SECONDS


def write_info(folder: Path, info: dict[str, str]) -> None:
    folder.mkdir(parents=True, exist_ok=True)
    (folder / "info.json").write_text(
        json.dumps(info, ensure_ascii=False, indent=2), encoding="utf-8"
    )


def _excel_row_key(promo, name, date_value, surgery) -> tuple[str, str, str, str]:
    """A~D열(홍보여부/이름/수술날짜/수술명)로 중복 여부를 판단하기 위한 키를
    만든다. 날짜 셀은 datetime으로 저장되므로 YYMMDD 문자열로 맞춰 비교한다."""
    if isinstance(date_value, datetime):
        date_key = date_value.strftime("%y%m%d")
    else:
        date_key = str(date_value).strip() if date_value else ""
    return (
        str(promo).strip() if promo else "",
        str(name).strip() if name else "",
        date_key,
        str(surgery).strip() if surgery else "",
    )


def append_excel_row(info: dict[str, str], excel_path: Path, log) -> None:
    """새 세션이 시작될 때 홍보 환자 엑셀에 한 행을 바로 추가한다."""
    try:
        import openpyxl
        from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    except ImportError as error:
        log(f"  엑셀 업데이트 건너뜀 (openpyxl 필요): {error}", "warn")
        return

    if not excel_path.exists():
        log(f"  엑셀 업데이트 건너뜀 (파일 없음): {excel_path}", "warn")
        return

    try:
        wb = openpyxl.load_workbook(excel_path)
        ws = wb.active

        existing_keys = set()
        for row in range(6, ws.max_row + 1):
            name_val = ws.cell(row=row, column=2).value
            if not name_val:
                continue
            existing_keys.add(_excel_row_key(
                ws.cell(row=row, column=1).value, name_val,
                ws.cell(row=row, column=3).value, ws.cell(row=row, column=4).value,
            ))

        new_key = _excel_row_key(info["홍보여부"], info["이름"], info["수술날짜"], info["수술명"])
        if new_key in existing_keys:
            log(f"  엑셀 업데이트 건너뜀 (이미 등록됨): {info['이름']}", "warn")
            return

        try:
            surgery_date = datetime.strptime(info["수술날짜"], "%y%m%d")
        except ValueError:
            surgery_date = None

        font_body = Font(name="Malgun Gothic", size=9)
        font_bold = Font(name="Malgun Gothic", size=9, bold=True)
        align_center = Alignment(horizontal="center", vertical="center", wrap_text=True)
        align_left = Alignment(horizontal="left", vertical="center", wrap_text=True)
        border_side = Side(border_style="thin", color="D3D3D3")
        border_cell = Border(left=border_side, right=border_side, top=border_side, bottom=border_side)
        fill_hp = PatternFill(start_color="FFE6CC", fill_type="solid")
        fill_ht = PatternFill(start_color="E2EFDA", fill_type="solid")
        fill_zebra = PatternFill(start_color="F0F4F8", fill_type="solid")

        row = ws.max_row + 1
        ws.row_dimensions[row].height = 22

        type_cell = ws.cell(row=row, column=1, value=info["홍보여부"])
        if info["홍보여부"] == "HP":
            type_cell.fill = fill_hp
        elif info["홍보여부"] == "HT":
            type_cell.fill = fill_ht

        ws.cell(row=row, column=2, value=info["이름"]).font = font_bold

        # 차트번호/전화번호였던 C, D열은 삭제되어 수술날짜가 C열로 당겨진다.
        date_cell = ws.cell(row=row, column=3)
        if surgery_date:
            date_cell.value = surgery_date
            date_cell.number_format = "yyyy-mm-dd"
        else:
            date_cell.value = ""

        ws.cell(row=row, column=4, value=info["수술명"])

        ws.cell(row=row, column=5, value=f"=C{row}+7").number_format = "yyyy-mm-dd"
        ws.cell(row=row, column=6, value="")
        ws.cell(row=row, column=7, value=f"=C{row}+14").number_format = "yyyy-mm-dd"
        ws.cell(row=row, column=8, value="")
        ws.cell(row=row, column=9,
                value=f"=DATE(YEAR(C{row}), MONTH(C{row})+1, DAY(C{row}))").number_format = "yyyy-mm-dd"
        ws.cell(row=row, column=10, value="")
        ws.cell(row=row, column=11,
                value=f"=DATE(YEAR(C{row}), MONTH(C{row})+3, DAY(C{row}))").number_format = "yyyy-mm-dd"
        ws.cell(row=row, column=12, value="")
        ws.cell(row=row, column=13,
                value=f"=DATE(YEAR(C{row}), MONTH(C{row})+6, DAY(C{row}))").number_format = "yyyy-mm-dd"
        ws.cell(row=row, column=14, value="")
        ws.cell(row=row, column=15, value="")

        is_even_row = (row % 2 == 0)
        for col in range(1, 16):
            cell = ws.cell(row=row, column=col)
            cell.border = border_cell
            if col != 2:
                cell.font = font_body
            cell.alignment = align_left if col in (4, 15) else align_center
            if col > 1 and is_even_row:
                cell.fill = fill_zebra

        wb.save(excel_path)
        log(f"  엑셀 업데이트: {info['이름']} 추가", "session")
    except Exception as error:
        log(f"  엑셀 업데이트 실패: {error}", "error")


def process_image(detector, image_path: Path, current_session: dict | None,
                   photo_dir: Path, backup_dir: Path, excel_path: Path, log,
                   use_pyzbar: bool = False) -> dict | None:
    payload = decode_qr(detector, image_path, use_pyzbar)

    if payload:
        info = parse_session_payload(payload)
        if info:
            parent_folder = build_session_parent_folder(info)
            session_folder = build_session_folder(info)
            backup_original(image_path, backup_dir, Path(parent_folder) / session_folder)
            marker_dir = photo_dir / "_세션마커" / parent_folder / session_folder
            safe_move(image_path, marker_dir)
            write_info(marker_dir, info)
            write_info(photo_dir / parent_folder / session_folder, info)
            if current_session and current_session["folder"] != session_folder:
                log(f"  이전 환자 종료: {current_session['folder']}", "session")
            log(f"▶ 새로운 환자 시작: {parent_folder}/{session_folder}", "session")
            append_excel_row(info, excel_path, log)
            return {"folder": session_folder, "info": info}

        backup_original(image_path, backup_dir)
        safe_move(image_path, photo_dir / "_미분류" / "알수없는_QR")
        log("  형식이 올바르지 않은 QR -> _미분류", "warn")
        return current_session

    if current_session:
        parent_folder = build_session_parent_folder(current_session["info"])
        subfolder = build_elapsed_subfolder(current_session["info"], image_path)
        target_dir = photo_dir / parent_folder / current_session["folder"] / subfolder
        target_dir.mkdir(parents=True, exist_ok=True)
        filename = next_sequence_filename(target_dir, subfolder, image_path.suffix)
        backup_original(
            image_path, backup_dir, Path(parent_folder) / current_session["folder"] / subfolder, filename
        )
        destination = target_dir / filename
        shutil.move(str(image_path), str(destination))
        log(f"  사진 저장: {parent_folder}/{current_session['folder']}/{subfolder}/{destination.name}", "photo")
        return current_session

    backup_original(image_path, backup_dir)
    safe_move(image_path, photo_dir / "_미분류")
    log("  시작 QR 전 사진 -> _미분류/", "warn")
    return None


def wait_for_directory(path: Path, stop_event: threading.Event, log, retry_seconds: float = 5.0) -> bool:
    """네트워크 드라이브(Z: 등)가 로그인 직후 아직 마운트되지 않았을 수 있어 재시도한다.
    중지 요청이 들어오면 False를 반환한다."""
    while not stop_event.is_set():
        try:
            path.mkdir(parents=True, exist_ok=True)
            return True
        except OSError as error:
            log(f"  {path} 접근 대기 중... ({error})", "warn")
            stop_event.wait(retry_seconds)
    return False


def run_watcher(watch_dir: Path, photo_dir: Path, backup_dir: Path, excel_path: Path,
                 stop_event: threading.Event, log_put, use_pyzbar: bool = False) -> None:
    """감시 루프 본체. 백그라운드 스레드에서 실행되며, log_put((level, str))으로 GUI에
    로그 한 줄씩 전달한다. level은 GUI에서 새 환자 시작(session)/사진 분류(photo)/
    미분류·대기(warn)/오류(error)/일반 안내(info)를 구분해 색을 다르게 표시하는 데 쓰인다."""

    def log(message: str, level: str = "info") -> None:
        log_put((level, f"{datetime.now().strftime('%H:%M:%S')} {message}"))

    try:
        import cv2
    except ImportError:
        log("OpenCV가 필요합니다. 아래 명령을 한 번 실행하세요:", "error")
        log("  .venv/Scripts/python.exe -m pip install opencv-python", "error")
        return

    if not wait_for_directory(watch_dir, stop_event, log):
        return
    if not wait_for_directory(photo_dir, stop_event, log):
        return

    current_session = load_current_session()

    log("=" * 50)
    log("QR 촬영 세션 자동 분류기 시작")
    log(f"감시 폴더: {watch_dir}")
    log(f"분류 대상 폴더: {photo_dir}")
    log(f"현재 진행 중인 환자: {current_session['folder'] if current_session else '없음'}")
    log("=" * 50)

    detector = cv2.QRCodeDetector()
    observations: dict[Path, tuple[int, int, float]] = {}

    while not stop_event.is_set():
        for image_path in incoming_images(watch_dir):
            if stop_event.is_set():
                break
            if not is_stable(image_path, observations):
                continue
            log(f"[처리] {image_path.name}")
            try:
                current_session = process_image(
                    detector, image_path, current_session, photo_dir, backup_dir, excel_path, log,
                    use_pyzbar,
                )
                save_current_session(current_session)
                observations.pop(image_path, None)
            except Exception as error:
                # 파일을 남겨 다음 주기에 다시 시도한다. 실패 파일을 처리 완료로 표시하지 않는다.
                log(f"  처리 보류 (다시 시도): {error}", "error")
        stop_event.wait(CHECK_INTERVAL)

    log("감시를 중지했습니다.")


# ==========================================
# 화면
# ==========================================
UI_FONT = ("Malgun Gothic", 10)
UI_FONT_BOLD = ("Malgun Gothic", 10, "bold")
TITLE_FONT = ("Malgun Gothic", 15, "bold")

BG_COLOR = "#f4f6fa"
CARD_COLOR = "#ffffff"
BORDER_COLOR = "#d7dbe3"
TEXT_COLOR = "#1f2937"
MUTED_COLOR = "#6b7280"
ACCENT_COLOR = "#2f6fed"
ACCENT_HOVER = "#255ac2"
SUCCESS_COLOR = "#1a7f37"
ERROR_COLOR = "#c0392b"
WARN_COLOR = "#b45309"

ENTRY_KWARGS = dict(
    font=UI_FONT, relief="flat", highlightthickness=1,
    highlightbackground=BORDER_COLOR, highlightcolor=ACCENT_COLOR,
    bg=CARD_COLOR, fg=TEXT_COLOR, insertbackground=TEXT_COLOR,
)

INITIAL_WINDOW_SIZE = (860, 760)


class QrClassifierApp(tk.Tk):
    def __init__(self) -> None:
        super().__init__()
        self.title("사진 자동 분류기")
        self.resizable(True, True)

        self._config = load_config()
        self._save_after_id = None
        self.watch_dir_var = tk.StringVar(value=self._config.get("watch_dir") or str(DEFAULT_WATCH_DIR))
        self.photo_dir_var = tk.StringVar(value=self._config.get("photo_dir") or str(DEFAULT_PHOTO_DIR))
        self.backup_dir_var = tk.StringVar(value=self._config.get("backup_dir") or str(DEFAULT_BACKUP_DIR))
        self.excel_path_var = tk.StringVar(value=self._config.get("excel_path") or str(DEFAULT_EXCEL_PATH))
        self.auto_start_var = tk.BooleanVar(value=self._config.get("auto_start", False))
        self.detailed_qr_var = tk.BooleanVar(value=self._config.get("detailed_qr_scan", False))
        for var in (self.watch_dir_var, self.photo_dir_var, self.backup_dir_var,
                    self.excel_path_var, self.auto_start_var, self.detailed_qr_var):
            var.trace_add("write", self._schedule_save_config)

        self._stop_event: threading.Event | None = None
        self._worker_thread: threading.Thread | None = None
        self._log_queue: Queue = Queue()

        self._apply_style()
        self._build_ui()
        self.minsize(*INITIAL_WINDOW_SIZE)
        self._center_window()
        self.protocol("WM_DELETE_WINDOW", self._on_close)

        # 예전 콘솔 버전처럼 실행하자마자 감시를 시작할지는 체크박스로 정한다
        # (로그인 자동 실행 시 사람이 버튼을 누를 필요가 없도록 기본은 켜둔다).
        if self.auto_start_var.get():
            self._start_watch()

    def _schedule_save_config(self, *_args) -> None:
        if self._save_after_id:
            self.after_cancel(self._save_after_id)
        self._save_after_id = self.after(500, self._save_config_now)

    def _save_config_now(self) -> None:
        self._save_after_id = None
        self._config.update({
            "watch_dir": self.watch_dir_var.get().strip(),
            "photo_dir": self.photo_dir_var.get().strip(),
            "backup_dir": self.backup_dir_var.get().strip(),
            "excel_path": self.excel_path_var.get().strip(),
            "auto_start": self.auto_start_var.get(),
            "detailed_qr_scan": self.detailed_qr_var.get(),
        })
        save_config(self._config)

    def _apply_style(self) -> None:
        self.configure(bg=BG_COLOR)
        self.option_add("*Font", UI_FONT)

        style = ttk.Style(self)
        style.theme_use("clam")
        style.configure(".", font=UI_FONT, background=BG_COLOR, foreground=TEXT_COLOR)
        style.configure("TFrame", background=BG_COLOR)
        style.configure("TLabel", background=BG_COLOR, foreground=TEXT_COLOR)
        style.configure("Header.TLabel", font=TITLE_FONT, background=BG_COLOR)
        style.configure("Muted.TLabel", background=BG_COLOR, foreground=MUTED_COLOR)
        style.configure("Success.TLabel", background=BG_COLOR, foreground=SUCCESS_COLOR)
        style.configure("Error.TLabel", background=BG_COLOR, foreground=ERROR_COLOR)

        style.configure("TLabelframe", background=BG_COLOR, bordercolor=BORDER_COLOR)
        style.configure("TLabelframe.Label", background=BG_COLOR, foreground=MUTED_COLOR,
                         font=UI_FONT_BOLD)

        style.configure("Accent.TButton", font=UI_FONT_BOLD, background=ACCENT_COLOR,
                         foreground="white", borderwidth=0, padding=(12, 10))
        style.map("Accent.TButton", background=[("active", ACCENT_HOVER), ("pressed", ACCENT_HOVER)])

        style.configure("Secondary.TButton", font=UI_FONT, background=CARD_COLOR,
                         foreground=TEXT_COLOR, bordercolor=BORDER_COLOR, padding=(10, 6))
        style.map("Secondary.TButton", background=[("active", "#eef1f6")])

        style.configure("TCheckbutton", background=BG_COLOR, foreground=TEXT_COLOR)
        style.map("TCheckbutton", background=[("active", BG_COLOR)])

    def _build_ui(self) -> None:
        self.columnconfigure(0, weight=1)
        self.rowconfigure(0, weight=1)

        outer = ttk.Frame(self, padding=24)
        outer.grid(row=0, column=0, sticky="nsew")
        outer.columnconfigure(0, weight=1)
        outer.rowconfigure(4, weight=1)

        ttk.Label(outer, text="사진 자동 분류기", style="Header.TLabel").grid(row=0, column=0, sticky="w")
        ttk.Label(outer, text="QR 촬영 세션에 맞춰 사진을 자동으로 분류합니다.",
                  style="Muted.TLabel").grid(row=1, column=0, sticky="w", pady=(2, 16))

        folders_frame = ttk.LabelFrame(outer, text="폴더 설정", padding=(16, 14))
        folders_frame.grid(row=2, column=0, sticky="ew")
        folders_frame.columnconfigure(1, weight=1)

        self._add_folder_row(folders_frame, 0, "감시 폴더", self.watch_dir_var)
        self._add_folder_row(folders_frame, 1, "분류 대상 폴더", self.photo_dir_var)
        self._add_folder_row(folders_frame, 2, "원본 백업 폴더", self.backup_dir_var)
        self._add_folder_row(folders_frame, 3, "엑셀 파일", self.excel_path_var,
                              browse_command=self._browse_excel_file,
                              extra_button=("열기", self._open_excel))

        control_frame = ttk.Frame(outer)
        control_frame.grid(row=3, column=0, sticky="ew", pady=(16, 8))

        self.toggle_btn = ttk.Button(control_frame, text="감시 시작", style="Accent.TButton",
                                      command=self._toggle_watch)
        self.toggle_btn.grid(row=0, column=0, sticky="w")

        self.status_var = tk.StringVar(value="중지됨")
        self.status_label = ttk.Label(control_frame, textvariable=self.status_var, style="Muted.TLabel")
        self.status_label.grid(row=0, column=1, sticky="w", padx=(16, 0))

        ttk.Checkbutton(control_frame, text="실행하면 자동으로 감시 시작", variable=self.auto_start_var,
                         style="TCheckbutton").grid(row=1, column=0, columnspan=2, sticky="w", pady=(8, 0))
        ttk.Checkbutton(control_frame, text="더 상세하게 QR 검사 (분류 속도가 느려집니다.)",
                         variable=self.detailed_qr_var,
                         style="TCheckbutton").grid(row=2, column=0, columnspan=2, sticky="w", pady=(4, 0))

        log_frame = ttk.LabelFrame(outer, text="실행 로그", padding=(10, 10))
        log_frame.grid(row=4, column=0, sticky="nsew", pady=(8, 0))
        log_frame.columnconfigure(0, weight=1)
        log_frame.rowconfigure(0, weight=1)

        self.log_text = scrolledtext.ScrolledText(
            log_frame, font=UI_FONT, wrap="word", state="disabled",
            bg=CARD_COLOR, fg=TEXT_COLOR, relief="flat", highlightthickness=1,
            highlightbackground=BORDER_COLOR,
        )
        self.log_text.grid(row=0, column=0, sticky="nsew")
        # 새 환자 시작(session)은 굵게 강조해 사진 분류(photo)와 한눈에 구분되게 한다.
        self.log_text.tag_configure("session", foreground=ACCENT_COLOR, font=UI_FONT_BOLD)
        self.log_text.tag_configure("photo", foreground=MUTED_COLOR)
        self.log_text.tag_configure("warn", foreground=WARN_COLOR)
        self.log_text.tag_configure("error", foreground=ERROR_COLOR, font=UI_FONT_BOLD)
        self.log_text.tag_configure("info", foreground=TEXT_COLOR)

    def _add_folder_row(self, parent: ttk.Frame, row: int, label: str, var: tk.StringVar,
                         browse_command=None, extra_button=None) -> None:
        ttk.Label(parent, text=label).grid(row=row, column=0, sticky="e", padx=(0, 12), pady=6)
        entry = tk.Entry(parent, textvariable=var, **ENTRY_KWARGS)
        entry.grid(row=row, column=1, sticky="ew", padx=(0, 8), pady=6)
        command = browse_command or (lambda v=var: self._browse_dir(v))
        ttk.Button(parent, text="찾아보기", style="Secondary.TButton",
                   command=command).grid(row=row, column=2, pady=6, padx=(0, 8) if extra_button else 0)
        if extra_button:
            text, extra_command = extra_button
            ttk.Button(parent, text=text, style="Secondary.TButton",
                       command=extra_command).grid(row=row, column=3, pady=6)

    def _browse_dir(self, var: tk.StringVar) -> None:
        initial = var.get().strip() or str(BASE_DIR)
        selected = filedialog.askdirectory(initialdir=initial, title="폴더 선택")
        if selected:
            var.set(selected)

    def _browse_excel_file(self) -> None:
        current = Path(self.excel_path_var.get().strip() or DEFAULT_EXCEL_PATH)
        selected = filedialog.askopenfilename(
            initialdir=str(current.parent) if current.parent.exists() else str(BASE_DIR),
            initialfile=current.name,
            title="엑셀 파일 선택",
            filetypes=[("Excel 파일", "*.xlsx"), ("모든 파일", "*.*")],
        )
        if selected:
            self.excel_path_var.set(selected)

    def _open_excel(self) -> None:
        excel_path = Path(self.excel_path_var.get().strip() or DEFAULT_EXCEL_PATH)
        if not excel_path.exists():
            messagebox.showerror("파일 없음", f"엑셀 파일을 찾을 수 없습니다:\n{excel_path}")
            return
        try:
            os.startfile(str(excel_path))
        except OSError as exc:
            messagebox.showerror("파일 열기 실패", str(exc))

    def _center_window(self) -> None:
        self.update_idletasks()
        width, height = INITIAL_WINDOW_SIZE
        x = (self.winfo_screenwidth() - width) // 2
        y = (self.winfo_screenheight() - height) // 3
        self.geometry(f"{width}x{height}+{x}+{y}")

    def _toggle_watch(self) -> None:
        if self._worker_thread and self._worker_thread.is_alive():
            self._stop_watch()
        else:
            self._start_watch()

    def _start_watch(self) -> None:
        watch_dir = self.watch_dir_var.get().strip()
        photo_dir = self.photo_dir_var.get().strip()
        backup_dir = self.backup_dir_var.get().strip()
        excel_path = self.excel_path_var.get().strip()
        if not (watch_dir and photo_dir and backup_dir and excel_path):
            messagebox.showerror(
                "입력 오류", "감시 폴더, 분류 대상 폴더, 원본 백업 폴더, 엑셀 파일을 모두 지정하세요."
            )
            return

        self._stop_event = threading.Event()
        self._log_queue = Queue()
        self._worker_thread = threading.Thread(
            target=run_watcher,
            args=(Path(watch_dir), Path(photo_dir), Path(backup_dir), Path(excel_path),
                  self._stop_event, self._log_queue.put, self.detailed_qr_var.get()),
            daemon=True,
        )
        self._worker_thread.start()
        self.toggle_btn.configure(text="감시 중지")
        self._set_status("감시 중", tone="success")
        self._poll_log_queue()

    def _stop_watch(self) -> None:
        if self._stop_event:
            self._stop_event.set()
        self.toggle_btn.configure(text="감시 시작")
        self._set_status("중지됨", tone="muted")

    def _set_status(self, text: str, *, tone: str = "muted") -> None:
        self.status_var.set(text)
        style_by_tone = {"muted": "Muted.TLabel", "success": "Success.TLabel", "error": "Error.TLabel"}
        self.status_label.configure(style=style_by_tone[tone])

    def _poll_log_queue(self) -> None:
        while True:
            try:
                level, message = self._log_queue.get_nowait()
            except Empty:
                break
            self._append_log(message, level)

        if self._worker_thread and self._worker_thread.is_alive():
            self.after(200, self._poll_log_queue)
        elif self.toggle_btn.cget("text") == "감시 중지":
            # 스레드가 오류 등으로 스스로 끝난 경우 버튼/상태 표시를 정리한다.
            self.toggle_btn.configure(text="감시 시작")
            self._set_status("중지됨", tone="muted")

    def _append_log(self, message: str, level: str = "info") -> None:
        self.log_text.configure(state="normal")
        self.log_text.insert("end", message + "\n", level)
        self.log_text.see("end")
        self.log_text.configure(state="disabled")

    def _on_close(self) -> None:
        if self._stop_event:
            self._stop_event.set()
        self.destroy()


def main() -> None:
    app = QrClassifierApp()
    app.mainloop()


if __name__ == "__main__":
    main()
