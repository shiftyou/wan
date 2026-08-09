import calendar
import os
import re
from datetime import datetime, timedelta
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.worksheet.datavalidation import DataValidation

# ==========================================
# [설정 부분] 병원 환경에 맞게 경로를 수정하세요
# ==========================================
# 1. 환자 사진들이 모여있는 폴더 경로 (QR로 생성되는 세션 폴더들이 위치하는 곳)
TARGET_DIR = os.environ.get("TARGET_DIR", r"Z:\01_환자이름별사진")

# 2. 업데이트할 엑셀 파일 경로
EXCEL_PATH = os.path.join(TARGET_DIR, "완성형_홍보환자_리스트.xlsx")

# ==========================================
# 폴더명 분석 함수
# ==========================================
def parse_folder_name(folder_name):
    """
    폴더명을 분석하여 (구분, 이름, 수술명, 수술날짜)를 반환합니다.
    우선순위 1: 언더바(_) 구분 규칙 -> HP_홍길동_쌍꺼풀수술_260429
    우선순위 2: 띄어쓰기 혼용 규칙 -> HP 홍길동_쌍꺼풀수술_260429
    첫 조각이 HP/HT/일반 중 하나면 그 값이 구분이 되고 그 다음 조각이 이름입니다.
    첫 조각이 셋 중 무엇도 아니면 구분은 기본값 '일반'이 되고, 첫 조각이 그대로 이름이 됩니다.
    구분을 뗀 나머지가 (이름, 수술명, 날짜) 최소 3조각으로 쪼개지지 않으면 None을 반환합니다.
    """
    # 1. 구분(HP, HT, 일반) 확인
    KNOWN_PREFIXES = ["HP", "HT", "일반"]
    prefix = "일반"
    clean_name = folder_name.strip()

    for p in KNOWN_PREFIXES:
        if clean_name.upper().startswith(p.upper() + " ") or clean_name.upper().startswith(p.upper() + "_"):
            prefix = p
            clean_name = clean_name[len(p) + 1:].strip()
            break

    # 2. 언더바(_) 기준으로 남은 부분 쪼개기
    parts = [p.strip() for p in clean_name.split("_") if p.strip()]
    
    # 예외 처리: 만약 언더바가 아예 없다면 띄어쓰기로 분리 시도
    if len(parts) < 2:
        parts = [p.strip() for p in clean_name.split(" ") if p.strip()]

    # 이름, 수술명, 날짜 최소 3개 조각이 없으면 대상에서 제외
    if len(parts) < 3:
        return None

    # 마지막 요소가 6자리 숫자(날짜)인지 확인 (예: 260429)
    date_str = None
    surgery_name = "정보 없음"
    patient_name = parts[0]

    last_part = parts[-1]
    # 6자리 숫자 패턴 검색
    if re.match(r"^\d{6}$", last_part):
        date_str = last_part
        # 중간에 낀 내용들을 모두 수술명으로 합침
        if len(parts) > 2:
            surgery_name = ", ".join(parts[1:-1])
    else:
        # 날짜 형식이 안 맞으면 전체를 수술명으로 처리
        surgery_name = ", ".join(parts[1:])

    return {
        "type": prefix,
        "name": patient_name,
        "surgery": surgery_name,
        "date_str": date_str
    }


# ==========================================
# 경과일 알림(빨간색 표시) 함수
# ==========================================
def add_months(base_date, months):
    """base_date에 개월 수를 더한 날짜를 반환한다 (말일 보정 포함)."""
    month_index = base_date.month - 1 + months
    year = base_date.year + month_index // 12
    month = month_index % 12 + 1
    day = min(base_date.day, calendar.monthrange(year, month)[1])
    return base_date.replace(year=year, month=month, day=day)


# (예정일 컬럼, 촬영여부 컬럼, 수술날짜로부터 예정일을 구하는 함수)
POD_RULES = [
    (7, 8, lambda d: d + timedelta(days=7)),    # G/H: POD #7
    (9, 10, lambda d: d + timedelta(days=14)),  # I/J: POD #14
    (11, 12, lambda d: add_months(d, 1)),       # K/L: POD #1M
    (13, 14, lambda d: add_months(d, 3)),       # M/N: POD #3M
    (15, 16, lambda d: add_months(d, 6)),       # O/P: POD #6M
]


def apply_pod_alerts(ws, fill_red, fill_zebra, fill_none):
    """예정일이 지났는데 촬영여부가 비어 있으면 예정일 셀 배경을 빨간색으로 바꾼다."""
    today = datetime.now().date()
    for row in range(6, ws.max_row + 1):
        surgery_val = ws.cell(row=row, column=5).value
        if not isinstance(surgery_val, datetime):
            continue
        surgery_date = surgery_val.date()
        is_even_row = (row % 2 == 0)
        default_fill = fill_zebra if is_even_row else fill_none

        for date_col, check_col, due_date_fn in POD_RULES:
            due_date = due_date_fn(surgery_date)
            check_val = ws.cell(row=row, column=check_col).value
            is_overdue = today >= due_date and (check_val is None or str(check_val).strip() == "")
            ws.cell(row=row, column=date_col).fill = fill_red if is_overdue else default_fill


# ==========================================
# 메인 실행 로직
# ==========================================
def main():
    if not os.path.exists(TARGET_DIR):
        print(f"[오류] 대상 폴더가 존재하지 않습니다: {TARGET_DIR}")
        return

    if not os.path.exists(EXCEL_PATH):
        print(f"[오류] 엑셀 파일이 존재하지 않습니다: {EXCEL_PATH}")
        return

    # 1. 대상 폴더에서 폴더 목록 읽기 및 분석
    print("폴더 스캔 중...")
    new_patients = []
    for item in os.listdir(TARGET_DIR):
        if item.startswith("_"):
            continue
        item_path = os.path.join(TARGET_DIR, item)
        if os.path.isdir(item_path):
            parsed = parse_folder_name(item)
            if parsed:
                new_patients.append(parsed)

    print(f"스캔 완료: 총 {len(new_patients)}명의 홍보 환자 후보 발견.")

    # 2. 기존 엑셀 파일 로드
    wb = openpyxl.load_workbook(EXCEL_PATH)
    ws = wb.active

    # 기존에 등록된 환자 식별용 데이터 수집 (중복 등록 방지)
    # 기준: 이름(B열) 만으로 중복 판단
    existing_records = set()
    for row in range(6, ws.max_row + 1):
        name_val = ws.cell(row=row, column=2).value
        if name_val:
            existing_records.add(str(name_val).strip())

    # 스타일 정의 (가독성을 위한 디자인 세팅)
    font_body = Font(name="Malgun Gothic", size=9)
    font_bold = Font(name="Malgun Gothic", size=9, bold=True)
    align_center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    align_left = Alignment(horizontal="left", vertical="center", wrap_text=True)
    
    border_side = Side(border_style="thin", color="D3D3D3")
    border_cell = Border(left=border_side, right=border_side, top=border_side, bottom=border_side)
    
    fill_hp = PatternFill(start_color="FFE6CC", fill_type="solid") # 살구색
    fill_ht = PatternFill(start_color="E2EFDA", fill_type="solid") # 연녹색
    fill_zebra = PatternFill(start_color="F0F4F8", fill_type="solid") # 홀수행 배경
    fill_red = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid") # 경과일 초과 경고
    fill_none = PatternFill(fill_type=None)

    # 3. 새로운 환자 등록 진행
    added_count = 0
    current_row = ws.max_row + 1

    for pat in new_patients:
        # 날짜 변환 (YYMMDD -> YYYY-MM-DD)
        date_obj = None
        if pat["date_str"]:
            try:
                # 2000년대 기준으로 변환
                date_obj = datetime.strptime(pat["date_str"], "%y%m%d")
            except ValueError:
                pass

        # 중복 체크 (이름 기준)
        if pat["name"] in existing_records:
            continue

        # 데이터 입력 시작
        ws.row_dimensions[current_row].height = 22

        # A열: 구분 (HP / HT / 일반)
        type_cell = ws.cell(row=current_row, column=1, value=pat["type"])
        if pat["type"] == "HP":
            type_cell.fill = fill_hp
        elif pat["type"] == "HT":
            type_cell.fill = fill_ht
        
        # B열: 이름
        ws.cell(row=current_row, column=2, value=pat["name"]).font = font_bold
        
        # C, D열: 차트번호, 전화번호 (공란으로 두고 나중에 수동 기입)
        ws.cell(row=current_row, column=3, value="")
        ws.cell(row=current_row, column=4, value="")
        
        # E열: 수술날짜
        date_cell = ws.cell(row=current_row, column=5)
        if date_obj:
            date_cell.value = date_obj
            date_cell.number_format = 'yyyy-mm-dd'
        else:
            date_cell.value = ""
            
        # F열: 수술명
        ws.cell(row=current_row, column=6, value=pat["surgery"])
        
        # G~P열: 경과일 계산 수식 및 촬영여부 세팅
        r = current_row
        # POD #7 예정일 및 촬영여부
        ws.cell(row=r, column=7, value=f"=E{r}+7").number_format = 'yyyy-mm-dd'
        ws.cell(row=r, column=8, value="")
        
        # POD #14 예정일 및 촬영여부
        ws.cell(row=r, column=9, value=f"=E{r}+14").number_format = 'yyyy-mm-dd'
        ws.cell(row=r, column=10, value="")
        
        # POD #1M 예정일 및 촬영여부
        ws.cell(row=r, column=11, value=f"=DATE(YEAR(E{r}), MONTH(E{r})+1, DAY(E{r}))").number_format = 'yyyy-mm-dd'
        ws.cell(row=r, column=12, value="")
        
        # POD #3M 예정일 및 촬영여부
        ws.cell(row=r, column=13, value=f"=DATE(YEAR(E{r}), MONTH(E{r})+3, DAY(E{r}))").number_format = 'yyyy-mm-dd'
        ws.cell(row=r, column=14, value="")
        
        # POD #6M 예정일 및 촬영여부
        ws.cell(row=r, column=15, value=f"=DATE(YEAR(E{r}), MONTH(E{r})+6, DAY(E{r}))").number_format = 'yyyy-mm-dd'
        ws.cell(row=r, column=16, value="")
        
        # Q열: 비고
        ws.cell(row=r, column=17, value="")

        # 스타일 일괄 적용
        is_even_row = (current_row % 2 == 0)
        for col in range(1, 18):
            cell = ws.cell(row=current_row, column=col)
            cell.border = border_cell
            
            # 폰트 설정
            if col != 2:
                cell.font = font_body
                
            # 정렬 설정
            if col in [6, 17]: # 수술명, 비고는 왼쪽 정렬
                cell.alignment = align_left
            else:
                cell.alignment = align_center
                
            # 구분(A) 열을 제외하고, 짝수/홀수 행에 따른 배경색 세팅 (얼룩말 패턴)
            if col > 1:
                if is_even_row:
                    cell.fill = fill_zebra

        current_row += 1
        added_count += 1
        # 기록 보관용 추가
        existing_records.add(pat["name"])

    # 4. 경과일 알림(빨간색) 갱신 - 신규 등록 여부와 무관하게 매번 확인
    apply_pod_alerts(ws, fill_red, fill_zebra, fill_none)

    # 5. 파일 저장
    wb.save(EXCEL_PATH)
    if added_count > 0:
        print(f"[성공] 총 {added_count}명의 새로운 홍보 환자가 엑셀에 추가되었습니다!")
    else:
        print("[안내] 새로 추가할 신규 홍보 환자 폴더가 없습니다.")
    print("[안내] 경과일 알림 색상 갱신 완료.")

if __name__ == "__main__":
    main()
